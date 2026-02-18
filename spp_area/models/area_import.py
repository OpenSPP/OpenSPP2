# Part of OpenSPP. See LICENSE file for full copyright and licensing details.
import base64
import datetime
import json
import logging
import math
import re
import time
from io import BytesIO

from openpyxl import load_workbook

from odoo import Command, _, api, fields, models
from odoo.exceptions import ValidationError

from odoo.addons.job_worker.delay import group

_logger = logging.getLogger(__name__)
_area_import_raw_model = "spp.area.import.raw"
_res_lang_model = "res.lang"
_area_import_channel = "area_import"

# Regex patterns for COD column name normalization
# Pattern 1: admin{N}Name_{lang} or admin{N}Pcode (Sri Lanka style)
_ADMIN_NAME_PATTERN = re.compile(r"^admin(\d+)Name_(\w+)$", re.IGNORECASE)
_ADMIN_PCODE_PATTERN = re.compile(r"^admin(\d+)Pcode$", re.IGNORECASE)
# Pattern 2: ADM{N}_{LANG} or ADM{N}_PCODE (standard COD style)
_ADM_PATTERN = re.compile(r"^ADM(\d+)_(\w+)$", re.IGNORECASE)


class OpenSPPAreaImport(models.Model):
    _name = "spp.area.import"
    _description = "Areas Import Table"
    _users_model = "res.users"

    JOB_QUEUE_BATCH_SIZE = 1000

    NEW = "New"
    UPLOADED = "Uploaded"
    PARSED = "Parsed"
    IMPORTED = "Imported"
    VALIDATED = "Validated"
    DONE = "Done"
    CANCELLED = "Cancelled"

    STATE_SELECTION = [
        (NEW, NEW),
        (UPLOADED, UPLOADED),
        (PARSED, PARSED),
        (IMPORTED, IMPORTED),
        (VALIDATED, VALIDATED),
        (DONE, DONE),
        (CANCELLED, CANCELLED),
    ]

    name = fields.Char("File Name", required=True, translate=True)
    excel_file = fields.Binary("Area Excel File")
    date_uploaded = fields.Datetime()

    upload_id = fields.Many2one(_users_model, "Uploaded by")
    date_parsed = fields.Datetime()
    parse_id = fields.Many2one(_users_model, "Parsed by")
    json_file_ids = fields.One2many("spp.area.import.json", "area_import_id", "JSON Files")
    date_imported = fields.Datetime()
    import_id = fields.Many2one(_users_model, "Imported by")
    date_validated = fields.Datetime()
    validate_id = fields.Many2one(_users_model, "Validated by")
    raw_data_ids = fields.One2many(_area_import_raw_model, "area_import_id", "Raw Data")
    total_rows_imported = fields.Integer(
        "Total Rows Imported",
        compute="_compute_get_total_rows",
        store=True,
        readonly=True,
    )
    total_rows_error = fields.Integer(
        "Total Rows with Error",
        compute="_compute_get_total_rows",
        store=True,
        readonly=True,
    )
    state = fields.Selection(
        STATE_SELECTION,
        "Status",
        default=NEW,
    )

    is_locked = fields.Boolean(default=False)
    locked_reason = fields.Char(readonly=True)

    # Progress tracking fields
    progress_total = fields.Integer("Total Tasks", default=0, readonly=True)
    progress_done = fields.Integer("Completed Tasks", default=0, readonly=True)
    progress_percent = fields.Float(
        "Progress (%)",
        compute="_compute_progress_percent",
        store=False,
    )

    @api.depends("progress_total", "progress_done")
    def _compute_progress_percent(self):
        for rec in self:
            if rec.progress_total > 0:
                rec.progress_percent = (rec.progress_done / rec.progress_total) * 100
            else:
                rec.progress_percent = 0

    def _reset_progress(self):
        """Reset progress tracking fields"""
        self.update(
            {
                "progress_total": 0,
                "progress_done": 0,
            }
        )

    def _set_progress(self, total, done=0):
        """Set progress tracking values"""
        self.update(
            {
                "progress_total": total,
                "progress_done": done,
            }
        )

    def _increment_progress(self):
        """Increment the progress counter by 1"""
        self.progress_done = self.progress_done + 1

    @api.onchange("excel_file")
    def excel_file_change(self):
        """
        The above function is an onchange function in Python that updates the date_uploaded, upload_id,
        and state fields based on the value of the excel_file field.
        """

        if self.name:
            self.update(
                {
                    "date_uploaded": fields.Datetime.now(),
                    "upload_id": self.env.user,
                    "state": self.UPLOADED,
                }
            )
        else:
            self.update({"date_uploaded": None, "upload_id": None, "state": self.UPLOADED})

    @api.depends("raw_data_ids", "raw_data_ids.state")
    def _compute_get_total_rows(self):
        """
        The function `_compute_get_total_rows` calculates the total number of imported rows and the
        total number of rows with an error for a given record.
        """
        for rec in self:
            tot_rows_imported = len(rec.raw_data_ids)
            tot_rows_error = self.env[_area_import_raw_model].search(
                [("id", "in", rec.raw_data_ids.ids), ("state", "=", "Error")]
            )
            rec.update(
                {
                    "total_rows_imported": tot_rows_imported,
                    "total_rows_error": len(tot_rows_error),
                }
            )

    def cancel_import(self):
        """
        The function cancels the import by updating the state of the record to "Cancelled".
        """
        for rec in self:
            rec.update({"state": self.CANCELLED})

    def reset_to_uploaded(self):
        """
        The function resets the state of a record to "Uploaded".
        """
        for rec in self:
            rec.update({"state": self.UPLOADED})

    def get_cell_value(self, sheet, row, col):
        # openpyxl worksheet
        # if isinstance(sheet, (Worksheet, ReadOnlyWorksheet)):
        #     return sheet.cell(row=row+1, column=col+1).value
        # else:
        #     # xlrd sheet
        return sheet.cell(row, col).value

    def _get_book(self):
        self.ensure_one()
        try:
            inputx = BytesIO()
            inputx.write(base64.decodebytes(self.excel_file))
        except TypeError as e:
            raise ValidationError(_("ERROR: {}").format(e)) from e

        filename = self.name.lower()
        if filename.endswith(".xlsx"):
            # Try to open with openpyxl first for .xlsx files
            try:
                book = load_workbook(inputx, read_only=True)
                return book
            except Exception as e:
                _logger.warning("Failed to open with openpyxl: %s", e)
        else:
            raise ValidationError(_("ERROR: Unsupported file format. Please upload a .xlsx file."))

    def get_sheet_openpyxl(self, book, name):
        return book[name]

    def get_columns_openpyxl(self, sheet):
        return [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]

    def _normalize_cod_header(self, header):
        """
        Normalize COD (Common Operational Dataset) column headers to a standard format.

        Handles different naming conventions found in HDX datasets:
        - Standard: ADM0_EN, ADM1_PCODE, ADM2_FR → kept as-is (uppercased)
        - Sri Lanka style: admin0Name_en, admin1Pcode → ADM0_EN, ADM1_PCODE

        Returns the normalized header string (uppercase).
        """
        if not header or not isinstance(header, str):
            return header

        # Try Sri Lanka style: admin{N}Name_{lang}
        match = _ADMIN_NAME_PATTERN.match(header)
        if match:
            level = match.group(1)
            lang = match.group(2).upper()
            return f"ADM{level}_{lang}"

        # Try Sri Lanka style: admin{N}Pcode
        match = _ADMIN_PCODE_PATTERN.match(header)
        if match:
            level = match.group(1)
            return f"ADM{level}_PCODE"

        # Standard ADM pattern - just uppercase it
        match = _ADM_PATTERN.match(header)
        if match:
            level = match.group(1)
            suffix = match.group(2).upper()
            return f"ADM{level}_{suffix}"

        # Other columns - keep as-is (preserve original case)
        return header

    def _detect_primary_language(self, headers):
        """
        Detect the primary language in the COD file by checking which language
        columns are present. Returns the 2-letter ISO code (e.g., 'EN', 'FR').

        Priority: EN > first available language found
        """
        languages_found = set()
        for header in headers:
            if not header:
                continue
            match = _ADM_PATTERN.match(header) or _ADMIN_NAME_PATTERN.match(header)
            if match:
                suffix = match.group(2).upper()
                # Skip non-language suffixes
                if suffix not in ("PCODE", "REF", "SQKM") and len(suffix) == 2:
                    languages_found.add(suffix)

        # Prefer English if available
        if "EN" in languages_found:
            return "EN"
        # Otherwise return first language found alphabetically
        if languages_found:
            return sorted(languages_found)[0]
        return "EN"  # Default fallback

    def process_file(self):
        """
        Process the uploaded Excel file in one operation: Parse → Import → Validate.
        This chains all preparation steps so users only need to review before Save to Area.

        After processing completes, users can review any validation errors in the Raw Data tab
        before clicking "Save to Area" to finalize the import.

        If missing languages are detected, shows a wizard to let user activate them first.
        """
        self.ensure_one()
        _logger.info("Area Import: Starting full file processing: %s", fields.Datetime.now())

        # Step 1: Parse Excel to JSON (synchronous - fast)
        self.parse_excel_to_json()

        # Step 2: Check for missing languages and show wizard if needed
        wizard_action = self._check_languages_and_show_wizard()
        if wizard_action:
            return wizard_action

        # Step 3: Import data (kicks off async jobs, will chain to validate when done)
        self._import_data_and_validate()

    def _continue_import_after_language_check(self):
        """
        Continue the import process after the language wizard has been handled.
        Called by the language wizard after user activates languages or chooses to skip.
        """
        self.ensure_one()
        _logger.info("Area Import: Continuing import after language check")

        # Continue with the import process
        self._import_data_and_validate()

        # Return action to reload the form
        return {
            "type": "ir.actions.act_window",
            "res_model": "spp.area.import",
            "res_id": self.id,
            "view_mode": "form",
            "target": "current",
        }

    def parse_excel_to_json(self):
        """
        Parse Excel file to JSON format synchronously.
        Uses openpyxl in read_only mode for memory-efficient parsing.
        Suitable for COD (Common Operational Dataset) files from HDX.

        Note: This operation runs synchronously because Excel parsing is fast
        (typically a few seconds) and doesn't require async queue processing.
        """
        self.ensure_one()
        _logger.info("Area Import: Starting Excel parse to JSON: %s", fields.Datetime.now())

        self.is_locked = True
        self.locked_reason = _("Parsing Excel file to JSON.")

        # Run parsing synchronously - it's fast enough and avoids queue worker dependency
        self._scan_and_create_parse_jobs()

    def _scan_and_create_parse_jobs(self):
        """
        Parse Excel file using openpyxl with read_only mode for memory efficiency.
        Creates JSON batches for each sheet, processing rows in batches.
        Suitable for COD (Common Operational Dataset) files from HDX.
        """
        self.ensure_one()
        parse_start = time.time()
        _logger.info("Area Import: Parsing Excel file with openpyxl at %s", fields.Datetime.now())

        # Clear existing JSON files
        if self.json_file_ids:
            self.json_file_ids.unlink()

        # Load Excel file with openpyxl in read_only mode for memory efficiency
        load_start = time.time()
        file_data = BytesIO(base64.decodebytes(self.excel_file))

        try:
            workbook = load_workbook(file_data, read_only=True, data_only=True)
        except Exception as e:
            _logger.error("Error reading Excel file with openpyxl: %s", e)
            raise ValidationError(_("Error reading Excel file: {}").format(str(e))) from e

        load_time = time.time() - load_start
        _logger.info("Area Import: Excel file loaded in %.2f seconds", load_time)

        # Track batches and rows
        batch_number = 0
        total_rows = 0
        total_batches = 0

        # Sort sheet names alphabetically to ensure consistent level assignment
        sheet_names = sorted(workbook.sheetnames)

        # Process each sheet
        for area_level, sheet_name in enumerate(sheet_names):
            sheet_start = time.time()
            sheet = workbook[sheet_name]

            _logger.info("Area Import: Processing sheet '%s' at level %s", sheet_name, area_level)

            # Read all rows from sheet (iter_rows is memory-efficient in read_only mode)
            all_rows = list(sheet.iter_rows(values_only=True))

            if not all_rows:
                _logger.info("Area Import: Sheet '%s' is empty, skipping", sheet_name)
                continue

            # First row is headers - filter out None values (trailing empty columns)
            raw_headers = all_rows[0]
            original_headers = [h for h in raw_headers if h is not None]
            header_count = len(original_headers)

            # Normalize headers to standard COD format (ADM{N}_{LANG})
            headers = [self._normalize_cod_header(h) for h in original_headers]

            # Detect primary language for this file (EN, FR, etc.)
            primary_language = self._detect_primary_language(headers)

            _logger.info(
                "Area Import: Sheet has %d rows and %d columns (primary language: %s)",
                len(all_rows) - 1,
                header_count,
                primary_language,
            )

            # Filter data rows (skip header, skip completely empty rows)
            data_rows = []
            for row in all_rows[1:]:
                # Trim row to actual header count and check if all values are None
                trimmed_row = row[:header_count]
                if any(cell is not None for cell in trimmed_row):
                    data_rows.append(trimmed_row)

            if not data_rows:
                _logger.info("Area Import: Sheet '%s' has no data rows, skipping", sheet_name)
                continue

            # Process in batches
            sheet_rows = 0
            batch_start = time.time()

            for start_idx in range(0, len(data_rows), self.JOB_QUEUE_BATCH_SIZE):
                end_idx = min(start_idx + self.JOB_QUEUE_BATCH_SIZE, len(data_rows))
                batch_rows = data_rows[start_idx:end_idx]

                # Convert batch to list of dicts (JSON-ready format)
                rows = []
                for row in batch_rows:
                    row_data = {
                        "_sheet_name": sheet_name,
                        "_area_level": area_level,
                        "_primary_language": primary_language,
                    }

                    # Add all columns with normalized headers
                    for col_idx, header in enumerate(headers):
                        value = row[col_idx] if col_idx < len(row) else None

                        # Handle datetime types for JSON serialization
                        if isinstance(value, datetime.datetime | datetime.date):
                            value = value.isoformat()

                        row_data[header] = value

                    rows.append(row_data)

                # Create JSON file for this batch
                if rows:
                    self._create_json_file(batch_number, rows, batch_start)
                    batch_number += 1
                    total_batches += 1
                    sheet_rows += len(rows)
                    total_rows += len(rows)
                    batch_start = time.time()

                # Log progress every 1000 rows
                if total_rows % 1000 == 0:
                    elapsed = time.time() - parse_start
                    rate = total_rows / elapsed if elapsed > 0 else 0
                    _logger.info(
                        "Area Import: Parsed %d rows in %.2fs (%.0f rows/s)",
                        total_rows,
                        elapsed,
                        rate,
                    )

            sheet_time = time.time() - sheet_start
            _logger.info(
                "Area Import: Completed sheet '%s' - %d rows in %.2f seconds",
                sheet_name,
                sheet_rows,
                sheet_time,
            )

        # Close workbook to free resources
        workbook.close()

        # Mark parsing as done
        total_time = time.time() - parse_start
        avg_time = total_time / total_rows if total_rows > 0 else 0
        rows_per_sec = total_rows / total_time if total_time > 0 else 0

        _logger.info(
            "Area Import: Completed parsing %d rows into %d JSON files "
            "in %.2f seconds (%.0f rows/s, avg %.4fs per row)",
            total_rows,
            total_batches,
            total_time,
            rows_per_sec,
            avg_time,
        )

        # Update state directly (no need for separate job)
        self.update(
            {
                "date_parsed": fields.Datetime.now(),
                "parse_id": self.env.user,
                "state": self.PARSED,
            }
        )

        self.is_locked = False
        self.locked_reason = None

    def _create_json_file(self, batch_number, rows, batch_start):
        """
        Create a JSON file from a batch of rows and store it as binary.
        """
        json_start = time.time()
        json_string = json.dumps(rows)
        json_bytes = json_string.encode("utf-8")
        json_time = time.time() - json_start

        batch_time = time.time() - batch_start

        _logger.info(
            "Area Import: Creating JSON file for batch %d with %d rows "
            "(batch time: %.2fs, JSON serialization: %.4fs, size: %d bytes)",
            batch_number,
            len(rows),
            batch_time,
            json_time,
            len(json_bytes),
        )

        # Create JSON file record
        self.env["spp.area.import.json"].create(
            {
                "area_import_id": self.id,
                "batch_number": batch_number,
                "json_file_name": f"batch_{batch_number}.json",
                "json_file": base64.b64encode(json_bytes),
                "row_count": len(rows),
                "file_size": len(json_bytes),
            }
        )

    def import_data(self):
        """
        Import data from parsed JSON files into raw data records.
        Each JSON file is processed as a separate batch job.
        Does NOT chain to validation - use process_file() for the full workflow.
        """
        self._do_import_data(chain_to_validate=False)

    def _import_data_and_validate(self):
        """
        Import data and automatically chain to validation when done.
        Used by process_file() for the unified workflow.
        """
        self._do_import_data(chain_to_validate=True)

    def _do_import_data(self, chain_to_validate=False):
        """
        Import data from parsed JSON files into raw data records.
        Each JSON file is processed as a separate batch job.

        Args:
            chain_to_validate: If True, automatically run validation after import completes.
        """
        self.ensure_one()
        _logger.info("Area Import: Started importing from JSON files: %s", fields.Datetime.now())

        if self.raw_data_ids:
            self.raw_data_ids.unlink()

        if not self.json_file_ids:
            raise ValidationError(_("No JSON files found. Please parse the Excel file first."))

        # Auto-activate required languages before importing
        self._activate_languages()

        # Set progress tracking
        total_batches = len(self.json_file_ids)
        self._set_progress(total_batches, 0)

        self.is_locked = True
        self.locked_reason = _("Importing data from JSON files.")
        jobs = []

        # Create a job for each JSON file batch
        for json_file in self.json_file_ids.sorted(lambda x: x.batch_number):
            jobs.append(self.delayable(channel=_area_import_channel)._import_data_from_json(json_file.id))

        main_job = group(*jobs)
        # Chain to validation if requested
        mark_done_func = "_import_and_validate_mark_done" if chain_to_validate else "_import_mark_done"
        main_job.on_done(self.delayable(channel=_area_import_channel)._async_mark_done(mark_done_func))
        main_job.delay()

    def _get_languages_from_json(self):
        """
        Extract language codes from the first JSON file.

        Returns:
            set: Set of language codes found (e.g., {'EN', 'AR', 'FR'})
        """
        self.ensure_one()

        if not self.json_file_ids:
            return set()

        # Get the first JSON file to check languages
        first_json_file = self.json_file_ids.sorted(lambda x: x.batch_number)[0]

        # Decode and parse JSON
        json_bytes = base64.b64decode(first_json_file.json_file)
        json_string = json_bytes.decode("utf-8")
        rows = json.loads(json_string)

        if not rows:
            return set()

        # Get first row to check available languages
        first_row = rows[0]

        # Find all language codes in the JSON
        found_languages = set()
        for key in first_row.keys():
            # Check for ADM pattern with language code (e.g., ADM0_EN, ADM1_FR)
            if key.startswith("ADM") and "_" in key:
                parts = key.split("_")
                if len(parts) >= 2:
                    # Last part should be language code (e.g., EN, FR, ES)
                    lang_code = parts[-1]
                    # Only consider 2-letter codes (avoid PCODE, SQKM, etc.)
                    if len(lang_code) == 2 and lang_code.isalpha():
                        found_languages.add(lang_code.upper())

        return found_languages

    def _activate_languages(self, language_codes=None):
        """
        Auto-activate languages that are needed for import.

        Args:
            language_codes: Set of language codes to activate. If None, uses languages from JSON.
        """
        self.ensure_one()

        if language_codes is None:
            language_codes = self._get_languages_from_json()

        if not language_codes:
            return

        _logger.info(
            "Area Import: Checking languages to activate: %s",
            ", ".join(sorted(language_codes)),
        )

        # Get active languages in Odoo
        active_languages = self.env[_res_lang_model].search([("active", "=", True)])
        active_iso_codes = set()
        for lang in active_languages:
            if lang.iso_code:
                active_iso_codes.add(lang.iso_code.upper())
            elif lang.code:
                active_iso_codes.add(lang.code.split("_")[0].upper())

        # Check for missing languages
        missing_languages = language_codes - active_iso_codes

        if not missing_languages:
            _logger.info("Area Import: All required languages are already active")
            return

        # Find and activate missing languages
        for lang_code in sorted(missing_languages):
            lang = (
                self.env[_res_lang_model]
                .with_context(active_test=False)
                .search([("iso_code", "=ilike", lang_code)], limit=1)
            )
            if not lang:
                lang = (
                    self.env[_res_lang_model]
                    .with_context(active_test=False)
                    .search([("code", "=ilike", f"{lang_code}_%")], limit=1)
                )
            if lang:
                _logger.info("Area Import: Auto-activating language: %s", lang.code)
                lang.active = True
            else:
                _logger.warning("Area Import: Language code %s not found in system", lang_code)

    def _validate_languages_activated(self):
        """
        Validate that all required languages are activated.
        Auto-activates missing languages if needed.
        """
        self.ensure_one()
        _logger.info("Area Import: Validating languages are activated...")

        # Auto-activate any missing languages
        self._activate_languages()

    def _check_languages_and_show_wizard(self):
        """
        Check if all languages found in JSON files are activated in Odoo.
        If missing languages are found, shows a wizard to let user choose which to activate.

        Returns:
            dict: Wizard action if missing languages found, None otherwise.
        """
        self.ensure_one()
        _logger.info("Area Import: Checking languages...")

        found_languages = self._get_languages_from_json()

        if not found_languages:
            return None

        _logger.info(
            "Area Import: Found languages in JSON: %s",
            ", ".join(sorted(found_languages)),
        )

        # Get active languages in Odoo
        # Build a set of ISO codes from active languages, handling missing iso_code gracefully
        active_languages = self.env[_res_lang_model].search([("active", "=", True)])
        active_iso_codes = set()
        for lang in active_languages:
            if lang.iso_code:
                active_iso_codes.add(lang.iso_code.upper())
            elif lang.code:
                # Fall back to extracting from code (e.g., "en_US" -> "EN")
                active_iso_codes.add(lang.code.split("_")[0].upper())

        _logger.info(
            "Area Import: Active language ISO codes: %s",
            ", ".join(sorted(active_iso_codes)),
        )

        # Check for missing languages
        missing_languages = found_languages - active_iso_codes

        if not missing_languages:
            return None  # All languages are already active

        # Find which languages can be activated (exist but inactive)
        activatable_langs = self.env[_res_lang_model]
        not_found = []

        for lang_code in sorted(missing_languages):
            # Search for the language (including inactive ones)
            # Try iso_code first, then fall back to code prefix (e.g., "en" matches "en_US")
            lang = (
                self.env[_res_lang_model]
                .with_context(active_test=False)
                .search([("iso_code", "=ilike", lang_code)], limit=1)
            )
            if not lang:
                # Try matching by code prefix (e.g., "en_US" starts with "en")
                lang = (
                    self.env[_res_lang_model]
                    .with_context(active_test=False)
                    .search([("code", "=ilike", f"{lang_code}_%")], limit=1)
                )
            if lang:
                activatable_langs |= lang
            else:
                not_found.append(lang_code)

        if not_found:
            _logger.warning(
                "Area Import: Unknown language codes will be skipped: %s",
                ", ".join(not_found),
            )

        # If no languages can be activated, just continue (unknown ones will be skipped)
        if not activatable_langs:
            return None

        _logger.info(
            "Area Import: Found %d activatable language(s): %s",
            len(activatable_langs),
            ", ".join(activatable_langs.mapped("name")),
        )

        # Create and return wizard action
        wizard = self.env["spp.area.import.language.wizard"].create(
            {
                "area_import_id": self.id,
                "language_ids": [Command.set(activatable_langs.ids)],
            }
        )

        return {
            "name": _("Activate Languages for Import"),
            "type": "ir.actions.act_window",
            "res_model": "spp.area.import.language.wizard",
            "res_id": wizard.id,
            "view_mode": "form",
            "target": "new",
        }

    def _import_data_from_json(self, json_file_id):
        """
        Import data from a single JSON file batch.
        Extracts area information based on the highest level in each row.
        Handles multi-language translations from Excel columns (ADM{level}_{LANG_CODE}).
        """
        self.ensure_one()
        import_start = time.time()

        json_file_record = self.env["spp.area.import.json"].browse(json_file_id)
        _logger.info(
            "Area Import: Processing %s (batch %d)",
            json_file_record.json_file_name,
            json_file_record.batch_number,
        )

        # Decode and parse JSON
        json_bytes = base64.b64decode(json_file_record.json_file)
        json_string = json_bytes.decode("utf-8")
        rows = json.loads(json_string)

        _logger.info(
            "Area Import: Loaded %d rows from %s",
            len(rows),
            json_file_record.json_file_name,
        )

        # Get active languages for mapping ISO codes
        active_languages = self.env[_res_lang_model].search([("active", "=", True)])
        lang_mapping = {lang.iso_code.upper(): lang.code for lang in active_languages}

        # Process each row
        for idx, row_data in enumerate(rows):
            row_start = time.time()

            area_level = row_data.get("_area_level", 0)
            primary_lang = row_data.get("_primary_language", "EN")

            # Extract current level fields (highest level in this row)
            admin_code_key = f"ADM{area_level}_PCODE"
            admin_code = row_data.get(admin_code_key)

            # Get default name using primary language (EN, FR, etc.)
            default_name = row_data.get(f"ADM{area_level}_{primary_lang}")

            # Find all translations for this level
            admin_level_prefix = f"ADM{area_level}_"
            # Map primary language to Odoo locale code
            primary_lang_code = lang_mapping.get(primary_lang, "en_US")
            translations = {primary_lang_code: default_name}

            for key, value in row_data.items():
                if key.startswith(admin_level_prefix) and key != admin_code_key:
                    lang_code = key.replace(admin_level_prefix, "")
                    # Skip non-language suffixes and the primary language
                    if lang_code in ("PCODE", "REF", "SQKM") or len(lang_code) != 2:
                        continue
                    if lang_code != primary_lang and lang_code in lang_mapping:
                        # If empty or None, use default name
                        if not value or (isinstance(value, str) and not value.strip()):
                            translations[lang_mapping[lang_code]] = default_name
                        else:
                            translations[lang_mapping[lang_code]] = value

            # Extract parent level fields (one level lower)
            parent_name = None
            parent_code = None
            if area_level > 0:
                parent_level = area_level - 1
                parent_name_key = f"ADM{parent_level}_{primary_lang}"
                parent_code_key = f"ADM{parent_level}_PCODE"
                parent_name = row_data.get(parent_name_key)
                parent_code = row_data.get(parent_code_key)

            # Create raw import record with default name
            raw_vals = {
                "area_import_id": self.id,
                "admin_name": default_name,
                "admin_code": admin_code,
                "parent_name": parent_name,
                "parent_code": parent_code,
                "level": area_level,
                "area_sqkm": row_data.get("AREA_SQKM"),
            }

            raw_record = self.env[_area_import_raw_model].create(raw_vals)

            # Update translations for all languages found
            for lang_code, translated_name in translations.items():
                # Skip the primary language (already set as default admin_name)
                if lang_code != primary_lang_code:
                    raw_record.with_context(lang=lang_code).write(
                        {
                            "admin_name": translated_name,
                        }
                    )

            row_time = time.time() - row_start
            if (idx + 1) % 10 == 0:  # Log every 10 rows
                _logger.info(
                    "Area Import: Processed %d/%d rows from batch %d (last row: %.4fs, translations: %d)",
                    idx + 1,
                    len(rows),
                    json_file_record.batch_number,
                    row_time,
                    len(translations),
                )

        batch_time = time.time() - import_start
        _logger.info(
            "Area Import: Completed batch %d - %d rows in %.2fs",
            json_file_record.batch_number,
            len(rows),
            batch_time,
        )

        # Update progress
        self._increment_progress()

    def _import_mark_done(self):
        """
        Mark the import as done after all batch jobs have completed.
        Called by _async_mark_done() after all JSON files have been processed.
        """
        self.ensure_one()

        # Update state to imported
        self.update(
            {
                "date_imported": fields.Datetime.now(),
                "import_id": self.env.user,
                "state": self.IMPORTED,
            }
        )

        _logger.info(
            "Area Import: All batches completed. Total raw records created: %s",
            len(self.raw_data_ids),
        )

    def _import_and_validate_mark_done(self):
        """
        Mark import as done and automatically chain to validation.
        Used by process_file() for the unified workflow.
        """
        self.ensure_one()

        # First mark import as done
        self.update(
            {
                "date_imported": fields.Datetime.now(),
                "import_id": self.env.user,
                "state": self.IMPORTED,
            }
        )

        _logger.info(
            "Area Import: Import completed. Total raw records: %s. Starting validation...",
            len(self.raw_data_ids),
        )

        # Chain to validation
        self._do_validate_raw_data()

    def validate_raw_data(self):
        """
        Validate raw data records - called from UI button.
        """
        for rec in self:
            rec._do_validate_raw_data()

    def _do_validate_raw_data(self):
        """
        Internal method to validate raw data records.
        Sets up async jobs to validate each batch and update state when done.
        """
        self.ensure_one()
        self.is_locked = True
        self.locked_reason = _("Validating data.")
        ceiling = self.JOB_QUEUE_BATCH_SIZE
        batches = math.ceil(len(self.raw_data_ids) / ceiling)

        # Set progress tracking
        self._set_progress(batches, 0)

        jobs = []
        for i in range(batches):
            start = i * ceiling
            end = min((i + 1) * ceiling, len(self.raw_data_ids))
            jobs.append(self.delayable(channel=_area_import_channel)._validate_raw_data(self.raw_data_ids[start:end]))
        main_job = group(*jobs)
        main_job.on_done(self.delayable(channel=_area_import_channel)._validate_mark_done())
        main_job.delay()

    def _validate_raw_data(self, raw_data_ids):
        """
        The function validates raw data and updates the state if there are no errors.
        """
        self.ensure_one()
        raw_data_ids.validate_raw_data()
        self._increment_progress()

    def _validate_mark_done(self):
        """
        Mark validation as done after all batch jobs have completed.
        Updates state to VALIDATED if no errors were found.
        """
        self.ensure_one()
        self.is_locked = False
        self.locked_reason = None

        # Check for validation errors
        error_count = self.env[_area_import_raw_model].search_count(
            [("id", "in", self.raw_data_ids.ids), ("state", "=", "Error")]
        )

        if error_count == 0:
            self.update(
                {
                    "date_validated": fields.Datetime.now(),
                    "validate_id": self.env.user,
                    "state": self.VALIDATED,
                }
            )
            _logger.info("Area Import: Validation completed successfully. Ready for review.")
        else:
            _logger.warning(
                "Area Import: Validation completed with %d errors. Please review the Raw Data tab.",
                error_count,
            )

    def fix_area_level_and_type(self):
        for rec in self:
            rec.is_locked = True
            rec.locked_reason = _("Fixing area level.")
            ceiling = self.JOB_QUEUE_BATCH_SIZE
            batches = math.ceil(len(rec.raw_data_ids) / ceiling)
            jobs = []
            for i in range(batches):
                start = i * ceiling
                end = min((i + 1) * ceiling, len(rec.raw_data_ids))
                jobs.append(
                    rec.delayable(channel=_area_import_channel)._fix_area_level_and_type(rec.raw_data_ids[start:end])
                )
            main_job = group(*jobs)
            main_job.on_done(rec.delayable(channel=_area_import_channel)._async_mark_done())
            main_job.delay()

    def _fix_area_level_and_type(self, raw_data_ids):
        """
        The function `fix_area_level_and_type` fixes the area level of the raw data.
        """
        self.ensure_one()
        raw_data_ids.fix_area_level_and_type()

    def _async_mark_done(self, function_mark_done=None):
        """
        The function `_async_mark_done` unlocks a resource by setting the `locked` attribute to `False`
        and clearing the `locked_reason` attribute.
        """
        self.ensure_one()

        self.is_locked = False
        self.locked_reason = None

        if function_mark_done:
            getattr(self, function_mark_done)()

    def save_to_area(self):
        """
        The function saves data to an area, either synchronously or asynchronously depending on the
        number of raw data records.
        """
        for rec in self:
            rec.is_locked = True
            rec.locked_reason = _("Saving to area.")

            # Set progress tracking
            ceiling = self.JOB_QUEUE_BATCH_SIZE
            total_batches = math.ceil(len(rec.raw_data_ids) / ceiling)
            rec._set_progress(total_batches, 0)

            rec._async_recursive_save_to_area(rec.raw_data_ids)

    def _async_recursive_save_to_area(self, raw_data_ids):
        """
        This is to ensure that the function `_save_to_area` is called recursively and in order until all raw data
        is saved to the area.
        """
        self.ensure_one()
        jobs = []
        ceiling = self.JOB_QUEUE_BATCH_SIZE
        jobs.append(self.delayable(channel=_area_import_channel)._save_to_area(raw_data_ids[:ceiling]))
        main_job = group(*jobs)
        count = len(raw_data_ids)
        if count <= ceiling:
            main_job.on_done(self.delayable(channel=_area_import_channel)._save_to_area_mark_done())
        else:
            main_job.on_done(
                self.delayable(channel=_area_import_channel)._async_recursive_save_to_area(raw_data_ids[ceiling:])
            )
        main_job.delay()

    def _save_to_area(self, raw_data_ids):
        """
        The function saves raw data to an area and updates the state to "DONE".
        """
        self.ensure_one()

        raw_data_ids.save_to_area()
        self._increment_progress()

    def _save_to_area_mark_done(self):
        self.ensure_one()
        self.is_locked = False
        self.locked_reason = None
        if not self.env[_area_import_raw_model].search(
            [("id", "in", self.raw_data_ids.ids), ("state", "=", "Validated")]
        ):
            self.update(
                {
                    "state": self.DONE,
                }
            )

    def refresh_page(self):
        """
        The function `refresh_page` returns a dictionary with the type and tag values to reload the
        page.
        :return: The code is returning a dictionary with two key-value pairs. The "type" key has the
        value "ir.actions.client" and the "tag" key has the value "reload".
        """
        return {
            "type": "ir.actions.client",
            "tag": "reload",
        }


# Assets Import Raw Data
class OpenSPPAreaImportActivities(models.Model):
    _name = _area_import_raw_model
    _description = "Area Import Raw Data"
    _order = "level"
    _area_model = "spp.area"

    NEW = "New"
    VALIDATED = "Validated"
    ERROR = "Error"
    UPDATED = "Updated"
    POSTED = "Posted"

    STATE_CHOICES = [
        (NEW, NEW),
        (VALIDATED, VALIDATED),
        (ERROR, ERROR),
        (UPDATED, UPDATED),
        (POSTED, POSTED),
    ]

    STATE_ORDER_STATE = {
        ERROR: 0,
        NEW: 1,
        VALIDATED: 2,
        UPDATED: 3,
        POSTED: 4,
    }

    area_import_id = fields.Many2one("spp.area.import", "Area Import", required=True)
    admin_name = fields.Char(translate=True)
    admin_code = fields.Char()

    parent_name = fields.Char()
    parent_code = fields.Char()

    level = fields.Integer()

    area_sqkm = fields.Char("Area (sq/km)")

    remarks = fields.Text("Remarks/Errors")
    state = fields.Selection(
        STATE_CHOICES,
        "Status",
        default="New",
    )
    state_order = fields.Integer(
        compute="_compute_state_order",
        store=True,
    )
    area_id = fields.Many2one(_area_model, "Area", readonly=True)

    @api.depends("state")
    def _compute_state_order(self):
        for rec in self:
            rec.state_order = self.STATE_ORDER_STATE[rec.state]

    def check_errors(self):
        self.ensure_one()
        errors = []
        if not self.admin_name or not self.admin_code:
            errors.append(_("Name and Code of area is required."))

        if self.area_sqkm:
            try:
                float(self.area_sqkm)
            except ValueError:
                errors.append(_("AREA_SQKM should be numerical."))

        if self.level == 0 and (self.parent_name or self.parent_code):
            errors.append(_("Level 0 area should not have a parent name and parent code."))

        if self.level != 0 and (not self.parent_name or not self.parent_code):
            errors.append(_("Level 1 and above area should have a parent name and parent code."))
        return errors

    def validate_raw_data(self):
        for rec in self:
            errors = rec.check_errors()

            if errors:
                state = self.ERROR
                remarks = "\n".join(errors)
            else:
                state = self.VALIDATED
                remarks = "No Error"

            rec.write(
                {
                    "remarks": remarks,
                    "state": state,
                }
            )

    def get_area_vals(self):
        self.ensure_one()

        parent_id = None
        if self.parent_name and self.parent_code:
            parent_id = (
                self.env[self._area_model]
                .search(
                    [
                        ("code", "=", self.parent_code),
                    ],
                    limit=1,
                )
                .id
            )

        area_sqkm = self.area_sqkm

        try:
            area_sqkm = float(area_sqkm)
        except ValueError:
            area_sqkm = 0.0

        return {
            "parent_id": parent_id,
            "draft_name": self.admin_name,
            "code": self.admin_code,
            "area_sqkm": area_sqkm,
            "area_type_id": self.env.ref("spp_area.admin_area_type").id,
        }

    def save_to_area(self):
        """
        The function saves data to the "spp.area" model in the database, updating existing records if
        they exist and creating new records if they don't.
        """
        active_languages = self.env[_res_lang_model].search([("active", "=", True)])
        for rec in self:
            area_vals = rec.get_area_vals()
            if area_id := self.env[self._area_model].search([("code", "=", rec.admin_code)]):
                state = self.UPDATED
                area_id.update(area_vals)
            else:
                state = self.POSTED
                area_id = self.env[self._area_model].create(area_vals)

            for lang in active_languages:
                area_id.with_context(lang=lang.code).write(
                    {
                        "draft_name": rec.with_context(lang=lang.code).admin_name,
                    }
                )
                # Commenting out the compute_name and compute_complete_name
                # to lessen the load on the server, as this will be called when the area is saved with draft_name
                # area_id.with_context(lang=lang.code)._compute_name()
                # area_id.with_context(lang=lang.code)._compute_complete_name()

            rec.update(
                {
                    "state": state,
                    "remarks": "Successfully save to Area",
                    "area_id": area_id.id,
                }
            )

    def fix_area_level_and_type(self):
        for rec in self:
            if rec.area_id and (rec.area_id.area_level != rec.level or not rec.area_id.area_type_id):
                parent_id = None
                if rec.parent_name and rec.parent_code:
                    parent_id = (
                        self.env[self._area_model]
                        .search(
                            [
                                ("code", "=", rec.parent_code),
                            ],
                            limit=1,
                        )
                        .id
                    )
                rec.area_id.update(
                    {
                        "area_type_id": rec.env.ref("spp_area.admin_area_type").id,
                        "parent_id": parent_id,
                    }
                )


# JSON File Storage Model
class OpenSPPAreaImportJSON(models.Model):
    _name = "spp.area.import.json"
    _description = "Area Import JSON Files"
    _order = "batch_number"

    area_import_id = fields.Many2one("spp.area.import", "Area Import", required=True, ondelete="cascade")
    batch_number = fields.Integer("Batch Number", required=True)
    json_file = fields.Binary("JSON File", required=True)
    json_file_name = fields.Char("JSON File Name")
    row_count = fields.Integer("Row Count", help="Number of rows in this JSON file")
    file_size = fields.Integer("File Size (bytes)", help="Size of the JSON file in bytes")
